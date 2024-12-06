from fastapi import APIRouter, Depends, Query, HTTPException
from fastapi.responses import StreamingResponse
from schemas.allMoList import Allpackagelistsresponse,Allbomresponse
from sqlmodel import Session, select
from database.base import get_session
from models.dbm_user import User
from models.dbm_mo_histories import AllPackageLists,AllBomLists
from api.users import get_current_user
from typing import Optional
import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
import io
import warnings
from sqlalchemy import literal
warnings.filterwarnings("ignore", category=UserWarning)
router = APIRouter()


@router.get("/package",response_model=Allpackagelistsresponse)
def read_package_list(item_name: Optional[str] = Query(None),
                     package: Optional[str] = Query(None),
                     bonding: Optional[str] = Query(None),
                     lot_code: Optional[str] = Query(None),
                     supply: Optional[str] = Query(None),
                     order_date_start: Optional[str] = Query(None),
                     order_date_end: Optional[str] = Query(None),
                     page: int = Query(1, ge=1),
                     page_size: int = Query(15, ge=1),
                     session: Session = Depends(get_session),
                     current_user:User = Depends(get_current_user)):
    query = select(AllPackageLists).where(AllPackageLists.item_code.like(literal('BC-%-AB')))
    if item_name:
        query = query.where(AllPackageLists.item_name.contains(item_name.upper()))  # 模糊匹配用户名
    if package:
        query = query.where(AllPackageLists.package.contains(package.upper()))
    if bonding:
        query = query.where(AllPackageLists.bonding.contains(bonding.upper()))
    if lot_code:
        query = query.where(AllPackageLists.lot_code.contains(lot_code.upper()))
    if supply:
        query = query.where(AllPackageLists.supply.contains(supply))
    if order_date_start:
        query = query.where(AllPackageLists.order_date >= order_date_start)
    if order_date_end:
        query = query.where(AllPackageLists.order_date <= order_date_end)

    query = query.order_by(AllPackageLists.order_date)

    total = len(session.exec(query).all())

    # 分页
    query = query.offset(page_size * (page - 1)).limit(page_size)

    results = session.exec(query).all()

    return {"code":0,"data":results,"total":total}

@router.get("/cp",response_model=Allpackagelistsresponse)
def read_cp_list(supply: Optional[str] = Query(None),
                     wafer_name: Optional[str] = Query(None),
                     lot_code: Optional[str] = Query(None),
                     program:Optional[str] = Query(None),
                     order_date_start: Optional[str] = Query(None),
                     order_date_end: Optional[str] = Query(None),
                     page: int = Query(1, ge=1),
                     page_size: int = Query(15, ge=1),
                     session: Session = Depends(get_session),
                     current_user:User = Depends(get_current_user)):
    query = select(AllPackageLists).where(AllPackageLists.item_code.like(literal('CL-%-CP')))
    if supply:
        query = query.where(AllPackageLists.supply.contains(supply.upper()))  # 模糊匹配用户名
    if wafer_name:
        query = query.where(AllPackageLists.item_name.contains(wafer_name.upper()))
    if lot_code:
        query = query.where(AllPackageLists.lot_code.contains(lot_code.upper()))
    if program:
        query = query.where(AllPackageLists.pgm_name.contains(program.upper()))
    if order_date_start:
        query = query.where(AllPackageLists.order_date >= order_date_start)
    if order_date_end:
        query = query.where(AllPackageLists.order_date <= order_date_end)

    query = query.order_by(AllPackageLists.order_date)

    total = len(session.exec(query).all())

    # 分页
    query = query.offset(page_size * (page - 1)).limit(page_size)

    results = session.exec(query).all()

    return {"code":0,"data":results,"total":total}

@router.get("/bom",response_model=Allbomresponse)
def read_package_bom(order_id: Optional[str] = Query(None),
                     session: Session = Depends(get_session),
                     current_user:User = Depends(get_current_user)):
    try:
        query = select(AllBomLists)
        if order_id:
            query = query.where(AllBomLists.order_id == order_id).order_by(AllBomLists.main_chip)
        results = session.exec(query).all()
        return {"code":0,"data":results}
    except Exception as e:
        raise HTTPException(status_code=404,detail=str(e))


@router.get("/package/queryDownload")
def get_query_package_list(item_name: Optional[str] = Query(None),
                     package: Optional[str] = Query(None),
                     lot_code: Optional[str] = Query(None),
                     bonding: Optional[str] = Query(None),
                     order_date_start: Optional[str] = Query(None),
                     order_date_end: Optional[str] = Query(None),
                     supply: Optional[str] = Query(None),
                     session: Session = Depends(get_session),
                     current_user:User = Depends(get_current_user)):
    query = select(AllPackageLists)
    if item_name:
        query = query.where(AllPackageLists.item_name.contains(item_name.upper()))  # 模糊匹配用户名
    if package:
        query = query.where(AllPackageLists.package.contains(package.upper()))
    if bonding:
        query = query.where(AllPackageLists.bonding.contains(bonding.upper()))
    if lot_code:
        query = query.where(AllPackageLists.lot_code.contains(lot_code.upper()))
    if supply:
        query = query.where(AllPackageLists.supply.contains(supply))
    if order_date_start:
        query = query.where(AllPackageLists.order_date >= order_date_start)
    if order_date_end:
        query = query.where(AllPackageLists.order_date <= order_date_end)

    query = query.order_by(AllPackageLists.order_date)
    results = session.exec(query).all()
    # 将查询结果转换为字典列表，方便后续使用pandas处理
    data = [r.model_dump() for r in results]

    # 使用pandas的DataFrame来处理数据
    df = pd.DataFrame(data)

    # 自定义列顺序
    custom_columns = {"order_id":"订单号",
                      "item_name":"芯片名称",
                      "package":"封装形式",
                      "lot_code":"打印批号",
                      "business_qty":"订单数量",
                      "arrive_qty":"到货数量",
                      "remark":"备注",
                      "order_date":"订单日期",
                      "assy_step":"加工方式",
                      "pgm_name":"程序名称",
                      "loading_method":"装片方式",
                      "bonding":"打线图",
                      "wire":"线材",
                      "package_remark":"特殊备注",
                      "complete_date":"结束日期",
                      "supply":"封装厂"}
    # 重新排列列顺序
    df = df[list(custom_columns.keys())]

    # 修改列名
    df.rename(columns=custom_columns, inplace=True)


    # 导出excel
    # 将 DataFrame 写入 Excel 文件
    excel_io = io.BytesIO()
    with pd.ExcelWriter(excel_io, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Package Data")

        workbook = writer.book
        worksheet = writer.sheets["Package Data"]

        # 设置表头样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4F81BD")
        alignment = Alignment(horizontal="left", vertical="center")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = alignment
            cell.border = border

        # 设置数据单元格样式
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
            for cell in row:
                cell.alignment = alignment
                cell.border = border

        # 固定列宽
        fixed_width = 15  # 固定宽度，单位为字符宽度
        for col_idx in range(1, len(df.columns) + 1):
            col_letter = get_column_letter(col_idx)
            worksheet.column_dimensions[col_letter].width = fixed_width

    excel_io.seek(0)  # 重置指针以便读取

    headers = {
        "Content-Disposition": "attachment; filename=package_data.xlsx",
        "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    }
    return StreamingResponse(excel_io, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers=headers)

    """
    # 导出csv
    # 转换为 CSV 格式，并修复中文问题
    csv_io = io.BytesIO()  # 使用二进制模式避免编码问题
    df.to_csv(csv_io, index=False, header=True, encoding="utf-8-sig")  # 使用 utf-8-sig 编码支持 Excel
    csv_io.seek(0)
    
    # 设置响应头，支持文件下载
    headers = {
        "Content-Disposition": "attachment; filename=package_data.xlsx",
        "Content-Type": "text/csv; charset=utf-8",
    }
    return StreamingResponse(csv_io, media_type="text/csv", headers=headers)
    """


@router.get("/package/selectDownload")
def get_select_package_lists(ids: str = Query(...),
                     session: Session = Depends(get_session),
                     current_user: User = Depends(get_current_user)):
    id_list = ids.split(",")
    query = select(AllPackageLists).where(AllPackageLists.id.in_(id_list))

    results = session.exec(query).all()
    # 将查询结果转换为字典列表，方便后续使用pandas处理
    data = [r.model_dump() for r in results]

    # 使用pandas的DataFrame来处理数据
    df = pd.DataFrame(data)

    # 自定义列顺序
    custom_columns = {"order_id": "订单号",
                      "item_name": "芯片名称",
                      "package": "封装形式",
                      "lot_code": "打印批号",
                      "business_qty": "订单数量",
                      "arrive_qty": "到货数量",
                      "remark": "备注",
                      "order_date": "订单日期",
                      "assy_step": "加工方式",
                      "pgm_name": "程序名称",
                      "loading_method": "装片方式",
                      "bonding": "打线图",
                      "wire": "线材",
                      "package_remark": "特殊备注",
                      "complete_date": "结束日期",
                      "supply": "封装厂"}
    # 重新排列列顺序
    df = df[list(custom_columns.keys())]

    # 修改列名
    df.rename(columns=custom_columns, inplace=True)

    # 导出excel
    # 将 DataFrame 写入 Excel 文件
    excel_io = io.BytesIO()
    with pd.ExcelWriter(excel_io, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Package Data")

        workbook = writer.book
        worksheet = writer.sheets["Package Data"]

        # 设置表头样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4F81BD")
        alignment = Alignment(horizontal="left", vertical="center")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = alignment
            cell.border = border

        # 设置数据单元格样式
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
            for cell in row:
                cell.alignment = alignment
                cell.border = border

        # 固定列宽
        fixed_width = 15  # 固定宽度，单位为字符宽度
        for col_idx in range(1, len(df.columns) + 1):
            col_letter = get_column_letter(col_idx)
            worksheet.column_dimensions[col_letter].width = fixed_width

    excel_io.seek(0)  # 重置指针以便读取

    headers = {
        "Content-Disposition": "attachment; filename=package_data.xlsx",
        "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    }
    return StreamingResponse(excel_io, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers=headers)


@router.get("/cp/queryDownload")
def get_query_cp_list(supply: Optional[str] = Query(None),
                      wafer_name: Optional[str] = Query(None),
                      lot_code: Optional[str] = Query(None),
                      program:Optional[str] = Query(None),
                      order_date_start: Optional[str] = Query(None),
                      order_date_end: Optional[str] = Query(None),
                      session: Session = Depends(get_session),
                      current_user: User = Depends(get_current_user)):
    query = select(AllPackageLists).where(AllPackageLists.item_code.like(literal('CL-%-CP')))
    if supply:
        query = query.where(AllPackageLists.supply.contains(supply.upper()))  # 模糊匹配用户名
    if wafer_name:
        query = query.where(AllPackageLists.item_name.contains(wafer_name.upper()))
    if lot_code:
        query = query.where(AllPackageLists.lot_code.contains(lot_code.upper()))
    if program:
        query = query.where(AllPackageLists.pgm_name.contains(program.upper()))
    if order_date_start:
        query = query.where(AllPackageLists.order_date >= order_date_start)
    if order_date_end:
        query = query.where(AllPackageLists.order_date <= order_date_end)

    query = query.order_by(AllPackageLists.order_date)
    results = session.exec(query).all()
    # 将查询结果转换为字典列表，方便后续使用pandas处理
    data = [r.model_dump() for r in results]

    # 使用pandas的DataFrame来处理数据
    df = pd.DataFrame(data)

    # 自定义列顺序
    custom_columns = {"order_id": "订单号",
                      "item_name": "晶圆名称",
                      "lot_code": "晶圆批号",
                      "business_qty": "订单数量",
                      "arrive_qty": "测完数量",
                      "cp_step": "测试流程",
                      "pgm_name": "程序名称",
                      "remark": "备注",
                      "complete_date": "结束日期",
                      "order_date": "订单日期",
                      "supply": "中测厂"}
    # 重新排列列顺序
    df = df[list(custom_columns.keys())]

    # 修改列名
    df.rename(columns=custom_columns, inplace=True)

    # 导出excel
    # 将 DataFrame 写入 Excel 文件
    excel_io = io.BytesIO()
    with pd.ExcelWriter(excel_io, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Cp Data")

        workbook = writer.book
        worksheet = writer.sheets["Cp Data"]

        # 设置表头样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4F81BD")
        alignment = Alignment(horizontal="left", vertical="center")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = alignment
            cell.border = border

        # 设置数据单元格样式
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
            for cell in row:
                cell.alignment = alignment
                cell.border = border

        # 固定列宽
        fixed_width = 15  # 固定宽度，单位为字符宽度
        for col_idx in range(1, len(df.columns) + 1):
            col_letter = get_column_letter(col_idx)
            worksheet.column_dimensions[col_letter].width = fixed_width

    excel_io.seek(0)  # 重置指针以便读取

    headers = {
        "Content-Disposition": "attachment; filename=package_data.xlsx",
        "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    }
    return StreamingResponse(excel_io, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers=headers)

    """
    # 导出csv
    # 转换为 CSV 格式，并修复中文问题
    csv_io = io.BytesIO()  # 使用二进制模式避免编码问题
    df.to_csv(csv_io, index=False, header=True, encoding="utf-8-sig")  # 使用 utf-8-sig 编码支持 Excel
    csv_io.seek(0)

    # 设置响应头，支持文件下载
    headers = {
        "Content-Disposition": "attachment; filename=package_data.xlsx",
        "Content-Type": "text/csv; charset=utf-8",
    }
    return StreamingResponse(csv_io, media_type="text/csv", headers=headers)
    """


@router.get("/cp/selectDownload")
def get_select_cp_lists(ids: str = Query(...),
                     session: Session = Depends(get_session),
                     current_user: User = Depends(get_current_user)):
    id_list = ids.split(",")
    query = select(AllPackageLists).where(AllPackageLists.id.in_(id_list))

    results = session.exec(query).all()
    # 将查询结果转换为字典列表，方便后续使用pandas处理
    data = [r.model_dump() for r in results]

    # 使用pandas的DataFrame来处理数据
    df = pd.DataFrame(data)

    # 自定义列顺序
    custom_columns = {"order_id": "订单号",
                      "item_name": "晶圆名称",
                      "lot_code": "晶圆批号",
                      "business_qty": "订单数量",
                      "arrive_qty": "测完数量",
                      "cp_step": "测试流程",
                      "pgm_name": "程序名称",
                      "remark": "备注",
                      "complete_date": "结束日期",
                      "order_date": "订单日期",
                      "supply": "中测厂"}
    # 重新排列列顺序
    df = df[list(custom_columns.keys())]

    # 修改列名
    df.rename(columns=custom_columns, inplace=True)

    # 导出excel
    # 将 DataFrame 写入 Excel 文件
    excel_io = io.BytesIO()
    with pd.ExcelWriter(excel_io, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Cp Data")
        workbook = writer.book
        worksheet = writer.sheets["Cp Data"]

        # 设置表头样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4F81BD")
        alignment = Alignment(horizontal="left", vertical="center")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = alignment
            cell.border = border

        # 设置数据单元格样式
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
            for cell in row:
                cell.alignment = alignment
                cell.border = border

        # 固定列宽
        fixed_width = 15  # 固定宽度，单位为字符宽度
        for col_idx in range(1, len(df.columns) + 1):
            col_letter = get_column_letter(col_idx)
            worksheet.column_dimensions[col_letter].width = fixed_width

    excel_io.seek(0)  # 重置指针以便读取

    headers = {
        "Content-Disposition": "attachment; filename=package_data.xlsx",
        "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    }
    return StreamingResponse(excel_io, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers=headers)